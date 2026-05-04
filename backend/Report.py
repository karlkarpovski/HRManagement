from flask import jsonify, make_response
from sqlalchemy import text
from db import engine_human, engine_payroll
import io

# Try to import export libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

def get_hr_report():
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT d.DepartmentName, COUNT(e.EmployeeID) as TotalEmployees
                FROM Departments d
                LEFT JOIN Employees e ON d.DepartmentID = e.DepartmentID
                GROUP BY d.DepartmentName
            """)
            result = conn.execute(query).mappings().all()
        return jsonify([dict(row) for row in result]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def get_payroll_report():
    try:
        with engine_payroll.connect() as p_conn:
            p_query = text("SELECT EmployeeID, SUM(NetSalary) as TotalPaid FROM salaries GROUP BY EmployeeID")
            p_result = p_conn.execute(p_query).mappings().all()
            p_data = {}
            for row in p_result:
                p_data[row['EmployeeID']] = float(row['TotalPaid'])

        with engine_human.connect() as h_conn:
            h_query = text("""
                SELECT e.EmployeeID, e.FullName, d.DepartmentName
                FROM Employees e
                LEFT JOIN Departments d ON e.DepartmentID = d.DepartmentID
            """)
            h_data = h_conn.execute(h_query).mappings().all()

        report = []
        for emp in h_data:
            dept = emp['DepartmentName'] if emp['DepartmentName'] else 'N/A'
            report.append({
                'EmployeeID': emp['EmployeeID'],
                'FullName': emp['FullName'],
                'DepartmentName': dept,
                'TotalEarnings': p_data.get(emp['EmployeeID'], 0)
            })
        return jsonify(report), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def get_dividend_report():
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT 
                    d.DividendID,
                    d.EmployeeID,
                    e.FullName,
                    p.PositionName,
                    r.RoleName,
                    d.DividendAmount,
                    d.DividendDate
                FROM Dividends d
                LEFT JOIN Employees e ON d.EmployeeID = e.EmployeeID
                LEFT JOIN Positions p ON e.PositionID = p.PositionID
                LEFT JOIN AUTH_DB.dbo.Users u ON e.EmployeeID = u.EmployeeID
                LEFT JOIN AUTH_DB.dbo.User_Role ur ON u.UserID = ur.UserID
                LEFT JOIN AUTH_DB.dbo.Roles r ON ur.RoleID = r.RoleID
                ORDER BY d.DividendDate DESC
            """)
            result = conn.execute(query).mappings().all()
        return jsonify([dict(row) for row in result]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_hr_report_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 501
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT d.DepartmentName, COUNT(e.EmployeeID) as TotalEmployees
                FROM Departments d
                LEFT JOIN Employees e ON d.DepartmentID = e.DepartmentID
                GROUP BY d.DepartmentName
            """)
            result = conn.execute(query).mappings().all()
            data = [dict(row) for row in result]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HR Report"
        ws.append(['Phong Ban', 'Tong So Nhan Vien'])
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row in data:
            ws.append([row['DepartmentName'], row['TotalEmployees']])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=hr_report.xlsx'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_hr_report_pdf():
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed"}), 501
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT d.DepartmentName, COUNT(e.EmployeeID) as TotalEmployees
                FROM Departments d
                LEFT JOIN Employees e ON d.DepartmentID = e.DepartmentID
                GROUP BY d.DepartmentName
            """)
            result = conn.execute(query).mappings().all()
            data = [dict(row) for row in result]
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph("Bao Cao Nhan Su - Nhan Vien Theo Phong Ban", styles['Title']))
        elements.append(Spacer(1, 0.5*inch))
        
        table_data = [['Phong Ban', 'Tong So Nhan Vien']]
        for row in data:
            table_data.append([row['DepartmentName'], str(row['TotalEmployees'])])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=hr_report.pdf'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_payroll_report_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 501
    try:
        with engine_payroll.connect() as p_conn:
            p_query = text("SELECT EmployeeID, SUM(NetSalary) as TotalPaid FROM salaries GROUP BY EmployeeID")
            p_result = p_conn.execute(p_query).mappings().all()
            p_data = {}
            for row in p_result:
                p_data[row['EmployeeID']] = float(row['TotalPaid'])
        
        with engine_human.connect() as h_conn:
            h_query = text("""
                SELECT e.EmployeeID, e.FullName, d.DepartmentName
                FROM Employees e
                LEFT JOIN Departments d ON e.DepartmentID = d.DepartmentID
            """)
            h_data = h_conn.execute(h_query).mappings().all()
        
        report = []
        for emp in h_data:
            dept = emp['DepartmentName'] if emp['DepartmentName'] else 'N/A'
            report.append({
                'EmployeeID': emp['EmployeeID'],
                'FullName': emp['FullName'],
                'DepartmentName': dept,
                'TotalEarnings': p_data.get(emp['EmployeeID'], 0)
            })
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Report"
        ws.append(['Ma Nhan Vien', 'Ho Ten', 'Phong Ban', 'Tong Thu Nhap'])
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row in report:
            ws.append([row['EmployeeID'], row['FullName'], row['DepartmentName'], row['TotalEarnings']])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=payroll_report.xlsx'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_payroll_report_pdf():
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed"}), 501
    try:
        with engine_payroll.connect() as p_conn:
            p_query = text("SELECT EmployeeID, SUM(NetSalary) as TotalPaid FROM salaries GROUP BY EmployeeID")
            p_result = p_conn.execute(p_query).mappings().all()
            p_data = {}
            for row in p_result:
                p_data[row['EmployeeID']] = float(row['TotalPaid'])
        
        with engine_human.connect() as h_conn:
            h_query = text("""
                SELECT e.EmployeeID, e.FullName, d.DepartmentName
                FROM Employees e
                LEFT JOIN Departments d ON e.DepartmentID = d.DepartmentID
            """)
            h_data = h_conn.execute(h_query).mappings().all()
        
        report = []
        for emp in h_data:
            dept = emp['DepartmentName'] if emp['DepartmentName'] else 'N/A'
            report.append({
                'EmployeeID': emp['EmployeeID'],
                'FullName': emp['FullName'],
                'DepartmentName': dept,
                'TotalEarnings': p_data.get(emp['EmployeeID'], 0)
            })
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph("Bao Cao Luong - Tong Thu Nhap Theo Nhan Vien", styles['Title']))
        elements.append(Spacer(1, 0.5*inch))
        
        table_data = [['Ma Nhan Vien', 'Ho Ten', 'Phong Ban', 'Tong Thu Nhap']]
        for row in report:
            table_data.append([str(row['EmployeeID']), row['FullName'], row['DepartmentName'], str(row['TotalEarnings'])])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=payroll_report.pdf'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_dividend_report_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 501
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT 
                    d.DividendID,
                    d.EmployeeID,
                    e.FullName,
                    p.PositionName,
                    r.RoleName,
                    d.DividendAmount,
                    d.DividendDate
                FROM Dividends d
                LEFT JOIN Employees e ON d.EmployeeID = e.EmployeeID
                LEFT JOIN Positions p ON e.PositionID = p.PositionID
                LEFT JOIN AUTH_DB.dbo.Users u ON e.EmployeeID = u.EmployeeID
                LEFT JOIN AUTH_DB.dbo.User_Role ur ON u.UserID = ur.UserID
                LEFT JOIN AUTH_DB.dbo.Roles r ON ur.RoleID = r.RoleID
                ORDER BY d.DividendDate DESC
            """)
            result = conn.execute(query).mappings().all()
            data = [dict(row) for row in result]
        
        if not data:
            return jsonify({"error": "No data"}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dividend Report"
        headers = list(data[0].keys())
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row in data:
            ws.append(list(row.values()))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=dividend_report.xlsx'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def export_dividend_report_pdf():
    if not REPORTLAB_AVAILABLE:
        return jsonify({"error": "reportlab not installed"}), 501
    try:
        with engine_human.connect() as conn:
            query = text("""
                SELECT 
                    d.DividendID,
                    d.EmployeeID,
                    e.FullName,
                    p.PositionName,
                    r.RoleName,
                    d.DividendAmount,
                    d.DividendDate
                FROM Dividends d
                LEFT JOIN Employees e ON d.EmployeeID = e.EmployeeID
                LEFT JOIN Positions p ON e.PositionID = p.PositionID
                LEFT JOIN AUTH_DB.dbo.Users u ON e.EmployeeID = u.EmployeeID
                LEFT JOIN AUTH_DB.dbo.User_Role ur ON u.UserID = ur.UserID
                LEFT JOIN AUTH_DB.dbo.Roles r ON ur.RoleID = r.RoleID
                ORDER BY d.DividendDate DESC
            """)
            result = conn.execute(query).mappings().all()
            data = [dict(row) for row in result]
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph("Lich Su Co Tuc", styles['Title']))
        elements.append(Spacer(1, 0.5*inch))
        
        if data:
            headers = list(data[0].keys())
            table_data = [headers]
            for row in data:
                table_data.append([str(v) for v in row.values()])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 12),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=dividend_report.pdf'
        return response, 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
